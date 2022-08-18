# -*- coding: utf-8 -*-
"""
Created on Tue Jul 12 22:24:15 2022

@author: coder
"""

import pandas as pd
import openpyxl as op
import smtplib, ssl
from email.message import EmailMessage

port = 465
email = 'example@gmail.com'
password = "password"

wb = op.load_workbook(r'C:\Users\coder\Documents\Testing.xlsx')
wa = wb.active
df = pd.read_excel(r'C:\Users\coder\Documents\Testing.xlsx')
email_index = df.columns.get_loc('Email')
reallist = []
for i in range(2, wa.max_row+1):
    #print("\n")
    #print("Row ", i, " data :")
    #print()
    list=[]
    for j in range(1, wa.max_column+1):
        cell_obj = wa.cell(row=i, column=j)
        #print(cell_obj.value, end=" ")
        x = wa.cell(row=i, column=j).value
        list.append(x)
        if j==(wa.max_column):
            reallist.append(list)
            
    print("sending " + str(list) +" to "+ str(wa.cell(row=i, column=email_index+ 1).value ))
    
    body_message = "sending " + str(list) +" to "+ str(wa.cell(row=i, column=email_index+ 1).value )
    receiver = str(wa.cell(row=i, column=email_index+ 1).value)
    context = ssl.create_default_context()
    server=smtplib.SMTP_SSL("smtp.gmail.com", port, context=context)
    server.login(email, password)
    em = EmailMessage()
    em['From'] = email
    em['To'] = receiver
    em['Subject'] = "your results"
    em.set_content(body_message)
    server.send_message(em)
    print("message sent")
print(reallist)
    #print("\n send "+ str(x))
    # for i in list:
    #     print(i, end=" ")
